from openpyxl import Workbook
from R4C.settings import BASE_DIR
from datetime import datetime


class ReportCreator:
    wb = Workbook()

    # Метод для создания Excel отчета
    @classmethod
    def create_report(cls, report_dict: dict):
        for model in tuple(report_dict.keys()):
            ws = cls.wb.create_sheet(model)
            ws['A1'] = 'Модель'
            ws['B1'] = 'Версия'
            ws['C1'] = 'Количество за неделю'
            counter = 2
            for version in report_dict[model]:
                ws[f'A{counter}'] = version.model
                ws[f'B{counter}'] = version.version
                ws[f'C{counter}'] = version.counter
                counter += 1
        if 'Sheet' in cls.wb.sheetnames:
            cls.wb.remove(cls.wb['Sheet'])
        suffix = str(datetime.now())[0:19]
        suffix = suffix.replace(":", "-").replace(" ", "_")
        filename = "report_" + suffix + ".xlsx"
        path = BASE_DIR + '/robots/reports/' + filename
        cls.wb.save(path)
        return path
